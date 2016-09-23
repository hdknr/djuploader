# -*- coding: utf-8 -*-
from openpyxl import load_workbook, Workbook
from openpyxl.writer.excel import save_virtual_workbook
from django.utils.encoding import force_unicode
from openpyxl.utils import get_column_letter


EXCEL2007 = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


class XlsxBaseReader(object):
    MIMETYPE = EXCEL2007

    def __init__(
        self, filename=None, headers=None, sheet=0, converter=None,
        data_only=True, skip=0
    ):
        '''
            :type header: dict or None
            :param haeder: if None, the first row is `header`
        '''
        self.converter = converter or (lambda a: a)
        self.filename = filename    # filename or file
        self.sheet = sheet
        self.skip = skip
        self.headers = headers
        if self.filename:
            self.book = load_workbook(
                filename=self.filename, data_only=data_only)

    def row_values(self, row):
        return [cell.value and self.converter(cell.value) or None
                for cell in row if cell]

    def __iter__(self):

        line = 0
        for row in self.book.worksheets[self.sheet].iter_rows(
                row_offset=self.skip):
            line += 1
            values = self.row_values(row)
            yield line, values


class XlsxReader(XlsxBaseReader):

    def __iter__(self):
        header = list(self.headers or [])
        line = -1
        for row in self.book.worksheets[self.sheet].rows:
            line += 1
            if line == 0 and len(header) == 0:
                header = [v and v.strip() for v in self.row_values(row)]
                continue
            values = self.row_values(row)
            rowdict = dict(zip(header, values))
            yield line, rowdict, []


class XlsxWriter(object):
    MIMETYPE = EXCEL2007

    def __init__(self, stream, src=None, start=1, **kwargs):
        '''
            :type header: dict or None
            :param haeder: if None, the first row is `header`
        '''
        self.stream = stream
        self.book = src and load_workbook(filename=src) or Workbook()
        self.sheet = self.book.active
        self.sheet.title = "NewSheet"
        self.row = start

    def writerow(self, row):
        for i in xrange(1, len(row) + 1):
            ci = "{0}{1}".format(get_column_letter(i), self.row)
            self.sheet.cell(ci).value = force_unicode(row[i - 1])
        self.row += 1

    def close(self):
        self.stream.write(save_virtual_workbook(self.book))
